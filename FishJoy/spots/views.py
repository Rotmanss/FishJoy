import io
import os
from datetime import datetime

from django.apps import apps
from django.contrib.auth import logout, login
from django.contrib.auth.views import LoginView
from django.core.files import File
from django.db.models import Count, F, IntegerField
from django.db.models.functions import Cast
from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse, HttpResponseNotFound, Http404
from django.urls import reverse_lazy
from django.utils.datastructures import MultiValueDictKeyError
from django.views.generic import ListView, DetailView, CreateView, FormView
from django.core import serializers as ser
from matplotlib import ticker

from rest_framework import generics, viewsets, mixins, status
from rest_framework.decorators import action
from rest_framework.permissions import IsAuthenticatedOrReadOnly
from rest_framework.response import Response

from FishJoy.settings import BASE_DIR
from .forms import *
from .models import *
from .permissions import IsOwnerOrAdmin
from .serializer import *
from .utils import DataMixin

import openpyxl as xl
import pandas as pd
import matplotlib.pyplot as plt


def pageNotFound(request, exception):
    return HttpResponseNotFound("<h1>Page was not found</h1>")


def about(request):
    mixin = DataMixin()
    context = mixin.get_user_context(title='About')
    return render(request, 'spots/about.html', context)


class FisherHome(DataMixin, ListView):
    model = Spots
    template_name = 'spots/home.html'
    context_object_name = 'spots'

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super().get_context_data(**kwargs)
        c_def = self.get_user_context(title='Home page')
        return {**context, **c_def}

    def get_queryset(self):
        return Spots.objects.order_by('-rating', 'pk').select_related('spot_category')


class ShowFish(DataMixin, ListView):
    model = Fish
    template_name = 'spots/show_fish.html'
    context_object_name = 'fish'

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super().get_context_data(**kwargs)
        c_def = self.get_user_context(title='Fish')
        return {**context, **c_def}

    # def get_queryset(self):
    #     return Spots.objects.filter(fish__slug=self.kwargs['single_fish_slug'])


class ShowBaits(DataMixin, ListView):
    model = Baits
    template_name = 'spots/show_baits.html'
    context_object_name = 'baits'

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super().get_context_data(**kwargs)
        c_def = self.get_user_context(title='Baits')
        return {**context, **c_def}


class ShowSpot(DataMixin, DetailView):
    model = Spots
    template_name = 'spots/show_spot.html'
    context_object_name = 'spot'
    slug_url_kwarg = 'spot_slug'

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super().get_context_data(**kwargs)
        c_def = self.get_user_context(title='Spot - ' + str(context['spot'].title))
        return {**context, **c_def}


class ShowSingleFish(DataMixin, DetailView):
    model = Fish
    template_name = 'spots/show_single_fish.html'
    context_object_name = 'fish'
    slug_url_kwarg = 'single_fish_slug'

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super().get_context_data(**kwargs)
        c_def = self.get_user_context(title='Fish - ' + str(context['fish'].name))
        return {**context, **c_def}


class ShowSpotsCategory(DataMixin, ListView):
    model = Spots
    template_name = 'spots/home.html'
    context_object_name = 'spots'

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super().get_context_data(**kwargs)
        c_def = self.get_user_context(title='Spot category - ' + str(context['spots'][0].spot_category))
        return {**context, **c_def}

    def get_queryset(self):
        return Spots.objects.filter(spot_category__slug=self.kwargs['spots_category_slug']).order_by('-rating').\
            select_related('spot_category')


class ShowFishCategory(DataMixin, ListView):
    model = Fish
    template_name = 'spots/show_fish.html'
    context_object_name = 'fish'

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super().get_context_data(**kwargs)
        c_def = self.get_user_context(title='Fish category - ' + str(context['fish'][0].fish_category))
        return {**context, **c_def}

    def get_queryset(self):
        return Fish.objects.filter(fish_category__slug=self.kwargs['fish_category_slug']).select_related(
            'fish_category')


class AddSpot(DataMixin, CreateView):
    form_class = AddSpotForm
    template_name = 'spots/add_form.html'
    success_url = reverse_lazy('home')

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super().get_context_data(**kwargs)
        c_def = self.get_user_context(title='Add spot')
        return {**context, **c_def}

    def form_valid(self, form):
        form.instance.user = self.request.user
        return super().form_valid(form)


class AddFish(DataMixin, CreateView):
    form_class = AddFishForm
    template_name = 'spots/add_form.html'
    success_url = reverse_lazy('fish')

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super().get_context_data(**kwargs)
        c_def = self.get_user_context(title='Add fish')
        return {**context, **c_def}

    def form_valid(self, form):
        form.instance.user = self.request.user
        return super().form_valid(form)


class AddBait(DataMixin, CreateView):
    form_class = AddBaitForm
    template_name = 'spots/add_form.html'
    success_url = reverse_lazy('baits')

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super().get_context_data(**kwargs)
        c_def = self.get_user_context(title='Add bait')
        return {**context, **c_def}

    def form_valid(self, form):
        form.instance.user = self.request.user
        return super().form_valid(form)


class RegisterUser(DataMixin, CreateView):
    form_class = RegisterUserForm
    template_name = 'spots/register.html'

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super().get_context_data(**kwargs)
        c_def = self.get_user_context(title='Register')
        return {**context, **c_def}

    def form_valid(self, form):
        user = form.save()
        login(self.request, user)
        return redirect('home')


class LoginUser(DataMixin, LoginView):
    form_class = LoginUserForm
    template_name = 'spots/login.html'
    success_url = reverse_lazy('home')

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super().get_context_data(**kwargs)
        c_def = self.get_user_context(title='Login')
        return {**context, **c_def}

    def get_success_url(self):
        return reverse_lazy('home')


def logout_user(request):
    logout(request)
    return redirect('login')


def search(request):
    mixin = DataMixin()
    context = mixin.get_user_context(title='Search')
    searched = request.POST['searched']

    spots_data = Spots.objects.filter(title__icontains=searched) | \
                 Spots.objects.filter(rating__icontains=searched) | \
                 Spots.objects.filter(location__icontains=searched) | \
                 Spots.objects.filter(max_depth__icontains=searched) | \
                 Spots.objects.filter(spot_category__name__icontains=searched)

    fish_data = Fish.objects.filter(name__icontains=searched) | \
                Fish.objects.filter(average_weight__icontains=searched) | \
                Fish.objects.filter(fish_category__name__icontains=searched)

    bait_data = Baits.objects.filter(name__icontains=searched) | \
                Baits.objects.filter(price__icontains=searched)

    if request.method == 'POST' and searched != '':
        context['data'] = [*spots_data, *fish_data, *bait_data]

    return render(request, 'spots/search.html', context)


def like_action(request, spot_slug):
    current_url = request.META.get('HTTP_REFERER')
    if request.method == 'POST' and request.user.is_authenticated:
        spot = get_object_or_404(Spots, slug=spot_slug)
        if request.user.username not in str(spot.liked_by.all()):
            spot.likes += 1
            spot.liked_by.add(request.user.id)
            spot.dislikes -= 1 if request.user.username in str(spot.disliked_by.all()) else 0
            spot.disliked_by.remove(request.user.id) if request.user.username in str(spot.disliked_by.all()) else None
            spot.rating = spot.calculate_rating()
            spot.save()
    return redirect(current_url)


def dislike_action(request, spot_slug):
    current_url = request.META.get('HTTP_REFERER')
    if request.method == 'POST' and request.user.is_authenticated:
        spot = get_object_or_404(Spots, slug=spot_slug)
        if request.user.username not in str(spot.disliked_by.all()):
            spot.dislikes += 1
            spot.disliked_by.add(request.user.id)
            spot.likes -= 1 if request.user.username in str(spot.liked_by.all()) else 0
            spot.liked_by.remove(request.user.id) if request.user.username in str(spot.liked_by.all()) else None
            spot.rating = spot.calculate_rating()
            spot.save()
    return redirect(current_url)


# Django REST Framework

class SpotsViewSet(viewsets.ModelViewSet):
    queryset = Spots.objects.all()
    serializer_class = SpotsSerializer
    permission_classes = (IsAuthenticatedOrReadOnly,)

    def get_permissions(self):
        if self.action in ['update', 'partial_update', 'destroy']:
            permission_classes = (IsAuthenticatedOrReadOnly, IsOwnerOrAdmin)
        else:
            permission_classes = (IsAuthenticatedOrReadOnly,)
        return (permission() for permission in permission_classes)

    @action(methods=['get'], detail=False)
    def categories(self, request):
        cats = SpotCategory.objects.all()
        return Response({'cats': [c.name for c in cats]})

    @action(methods=['get'], detail=True)
    def category(self, request, pk=None):
        cat = SpotCategory.objects.get(pk=pk)
        return Response({'cat': cat.name})


class FishViewSet(viewsets.ModelViewSet):
    queryset = Fish.objects.all()
    serializer_class = FishSerializer
    permission_classes = (IsAuthenticatedOrReadOnly,)

    def get_permissions(self):
        if self.action in ['update', 'partial_update', 'destroy']:
            permission_classes = (IsAuthenticatedOrReadOnly, IsOwnerOrAdmin)
        else:
            permission_classes = (IsAuthenticatedOrReadOnly,)
        return (permission() for permission in permission_classes)

    @action(methods=['get'], detail=False)
    def categories(self, request):
        cats = FishCategory.objects.all()
        return Response({'cats': [c.name for c in cats]})

    @action(methods=['get'], detail=True)
    def category(self, request, pk=None):
        cat = FishCategory.objects.get(pk=pk)
        return Response({'cat': cat.name})


class BaitsViewSet(viewsets.ModelViewSet):
    queryset = Baits.objects.all()
    serializer_class = BaitsSerializer
    permission_classes = (IsAuthenticatedOrReadOnly,)

    def get_permissions(self):
        if self.action in ['update', 'partial_update', 'destroy']:
            permission_classes = (IsAuthenticatedOrReadOnly, IsOwnerOrAdmin)
        else:
            permission_classes = (IsAuthenticatedOrReadOnly,)
        return (permission() for permission in permission_classes)


class Feedback(DataMixin, CreateView):
    form_class = FeedbackForm
    template_name = 'spots/add_form.html'
    success_url = reverse_lazy('home')

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super().get_context_data(**kwargs)
        c_def = self.get_user_context(title='Feedback')
        return {**context, **c_def}


def spots_xml_export(request):
    queryset = Spots.objects.all()
    xml_data = ser.serialize('xml', queryset, indent=4)

    response = HttpResponse(xml_data, content_type='text/xml')
    response['Content-Disposition'] = 'attachment; filename="spots.xml"'
    return response


def fish_xml_export(request):
    queryset = Fish.objects.all()
    xml_data = ser.serialize('xml', queryset, indent=4)

    response = HttpResponse(xml_data, content_type='text/xml')
    response['Content-Disposition'] = 'attachment; filename="fish.xml"'
    return response


def baits_xml_export(request):
    queryset = Baits.objects.all()
    xml_data = ser.serialize('xml', queryset, indent=4)

    response = HttpResponse(xml_data, content_type='text/xml')
    response['Content-Disposition'] = 'attachment; filename="baits.xml"'
    return response


def spots_xlsx_export(request):
    spots = Spots.objects.all()

    wb = xl.Workbook()
    ws = wb.active

    ws.append(['pk', 'title', 'slug', 'rating', 'location', 'photo', 'max_depth', 'fish', 'spot_category', 'time_create', 'time_update', 'likes', 'dislikes',
              'user', 'liked_by', 'disliked_by'])

    for spot in spots:
        fish = ", ".join([str(f) for f in spot.fish.all()])
        liked_by = ", ".join([str(l) for l in spot.liked_by.all()])
        disliked_by = ", ".join([str(d) for d in spot.disliked_by.all()])
        ws.append([spot.pk, spot.title, spot.slug, spot.rating, spot.location, str(spot.photo), spot.max_depth, fish, str(spot.spot_category), str(spot.time_create), str(spot.time_update),
                   spot.likes, spot.dislikes, str(spot.user), liked_by, disliked_by])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="spots.xlsx"'
    wb.save(response)

    return response


def fish_xlsx_export(request):
    fish = Fish.objects.all()

    wb = xl.Workbook()
    ws = wb.active

    ws.append(['pk', 'name', 'slug', 'photo', 'average_weight', 'baits', 'fish_category', 'time_create', 'time_update', 'user'])

    for f in fish:
        bait = ", ".join([str(b) for b in f.baits.all()])
        ws.append([f.pk, f.name, f.slug, str(f.photo), f.average_weight, bait, str(f.fish_category), str(f.time_create), str(f.time_update),
                   str(f.user)])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="fish.xlsx"'
    wb.save(response)

    return response


def baits_xlsx_export(request):
    baits = Baits.objects.all()

    wb = xl.Workbook()
    ws = wb.active

    ws.append(['pk', 'name', 'slug', 'photo', 'price', 'time_create', 'time_update', 'user'])

    for bait in baits:
        ws.append([bait.pk, bait.name, bait.slug, str(bait.photo), bait.price, str(bait.time_create), str(bait.time_update), str(bait.user)])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="baits.xlsx"'
    wb.save(response)

    return response


def create_spots_diagram(request):
    mixin = DataMixin()
    context = mixin.get_user_context(title='Spots diagram')

    spots = Spots.objects.values('spot_category').annotate(count=Count('id'))
    spots = spots.annotate(id_count=Cast('count', output_field=IntegerField()))

    df = pd.DataFrame(list(spots), columns=['spot_category', 'id_count'])
    df = df.sort_values(by='spot_category', ascending=True)

    category_ids = list(df['spot_category'])
    categories = SpotCategory.objects.filter(id__in=category_ids).order_by('id').values_list('name', flat=True)
    plt.xlabel('Category')
    plt.ylabel('Number of records')
    plt.title('Number of records per category')
    plt.yticks(range(0, int(df['id_count'].max())+1))

    plt.bar(categories, df['id_count'])

    directory = os.path.join(BASE_DIR, 'spots/static/spots/images')
    plt.savefig(os.path.join(directory, 'spots_bar.png'))

    context['spots_diagram'] = '/static/spots/images/spots_bar.png'

    return render(request, 'spots/diagram.html', context=context)


def create_fish_diagram(request):
    mixin = DataMixin()
    context = mixin.get_user_context(title='Fish diagram')

    fish = Fish.objects.values('fish_category').annotate(Count('id'))
    df = pd.DataFrame(list(fish), columns=['fish_category', 'id__count'])
    df = df.sort_values(by='fish_category', ascending=True)

    category_ids = list(df['fish_category'])
    categories = FishCategory.objects.filter(id__in=category_ids).order_by('id').values_list('name', flat=True)

    plt.xlabel('Category')
    plt.ylabel('Number of records')
    plt.title('Number of records per category')
    plt.yticks(range(0, int(df['id__count'].max())+1))
    plt.bar(categories, df['id__count'])

    directory = os.path.join(BASE_DIR, 'spots/static/spots/images')
    plt.savefig(os.path.join(directory, 'fish_bar.png'))

    context['fish_diagram'] = '/static/spots/images/fish_bar.png'

    return render(request, 'spots/diagram.html', context=context)
